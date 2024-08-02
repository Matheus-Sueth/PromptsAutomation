from genesys.api import Genesys
import openpyxl
import openpyxl.workbook

api_genesys = Genesys('MOVIDA')
workbook = openpyxl.load_workbook('./docs/dados.xlsx')
workbook_2 = openpyxl.Workbook()
ws = workbook_2.active
ws.append(['Prompt','PromptId','Encontrado'])

for sheetname in workbook.sheetnames:
    sheet = workbook[sheetname]
    sucesso = falha = 0
    for prompt, conteudo in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        print(prompt)
        try:
            if not prompt is None:
                user_prompt = api_genesys.get_user_prompt_by_name_or_description(name=prompt)
                ws.append([user_prompt.name,user_prompt.id, True])
                api_genesys.pr
                sucesso += 1
        except Exception as error:
            ws.append([prompt,'', False])
            falha += 1
        print(f'{sucesso=}---{falha=}')
    workbook_2.save(f'./docs/prompts2.xlsx')
    