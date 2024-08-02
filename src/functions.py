import os
import shutil
from openpyxl import Workbook, load_workbook
from genesys.api import Genesys


def upload_user_prompts_for_ura(api_genesys: Genesys, diretorio: str, ivr: str, type_prompt: str, qtd_linhas_dataframe: int, language:str = 'pt-br', prefixo_prompt: str = '', sufixo_prompt: str = ''):
    """
    Criação de novos prompts, por um excel
    """
    lista_prompts = os.listdir(diretorio)
    lista_prompts_divididos: list[range] = [lista_prompts[l:l+qtd_linhas_dataframe] for l in range(0, len(lista_prompts), qtd_linhas_dataframe)]

    for range_prompt in lista_prompts_divididos:
        pasta_principal = ivr[4:] if 'IVR-' in ivr else ivr
        pasta_secundaria = f'{pasta_principal}/{range_prompt.stop}/'
        os.mkdir(pasta_secundaria)
        wb = Workbook()
        ws = wb.active
        ws.append(['Nome', 'Tipo', 'IVR', 'Prompt', 'Sucesso', 'Message'])

        for prompt in range_prompt:
            try:
                prompt_path = os.path.join(diretorio,prompt)
                shutil.copy(prompt_path, os.path.join(pasta_secundaria,prompt))
                prompt_name = prefixo_prompt + prompt[0:prompt.index(".")] + sufixo_prompt
                user_prompt = api_genesys.post_architect_prompts(prompt_name, ivr)
                user_prompt_resource = api_genesys.post_architect_prompt_resources(user_prompt.id, language)
                api_genesys.post_architect_prompt_upload(user_prompt_resource.upload_uri, prompt_name, prompt_path)
                success, msg = True, 'Prompt gravado com sucesso'
            except Exception as erro:
                success, msg = False, erro
            finally:
                ws.append([prompt_name, type_prompt, ivr, prompt, success, msg])

        wb.save(f'{pasta_secundaria}/dados.xlsx')


def upload_user_prompts_for_ura2(api_genesys: Genesys, diretorio: str, ivr: str, language:str = 'pt-br', prefixo_prompt: str = '', sufixo_prompt: str = ''):
    """
    Criação de novos prompts, pelos arquivos de prompts no diretorio
    """
    lista_prompts = os.listdir(diretorio)
    wb = Workbook()
    ws = wb.active
    ws.append(['Nome', 'IVR', 'Prompt', 'Sucesso', 'Message'])
    ivr_pasta = f'{diretorio}/{ivr[4:]}/'
    os.mkdir(ivr_pasta)
    for prompt in lista_prompts:
        prompt_path = os.path.join(diretorio,prompt)
        if not os.path.isfile(prompt_path):
            continue
        try:
            new_dir = os.path.join(ivr_pasta,prompt)
            shutil.copy(prompt_path, new_dir)
            os.remove(prompt_path)
            prompt_name = prefixo_prompt + prompt[0:prompt.index(".")] + sufixo_prompt
            user_prompt = api_genesys.post_architect_prompts(prompt_name, ivr)
            user_prompt_resource = api_genesys.post_architect_prompt_resources(user_prompt.id, language)
            api_genesys.post_architect_prompt_upload(user_prompt_resource.upload_uri, prompt_name  + '.wav', new_dir)
            success, msg = True, 'Prompt gravado com sucesso'
        except Exception as erro:
            success, msg = False, str(erro)
        finally:
            ws.append([prompt_name, ivr, prompt, success, msg])

    wb.save(f'{ivr_pasta}/dados.xlsx')


def upload_user_prompts_for_ura3(api_genesys: Genesys, diretorio: str, ivr: str, language:str = 'pt-br', prefixo_prompt: str = '', sufixo_prompt: str = ''):
    """
    Atualização de prompts, pelos arquivos de prompts no diretorio
    """
    lista_prompts = os.listdir(diretorio)
    wb = Workbook()
    ws = wb.active
    ws.append(['Nome', 'IVR', 'Prompt', 'Sucesso', 'Message'])
    ivr_pasta = f'{diretorio}/{ivr[4:]}/'
    os.mkdir(ivr_pasta)
    for prompt in lista_prompts:
        prompt_path = os.path.join(diretorio,prompt)
        if not os.path.isfile(prompt_path):
            continue
        try:
            new_dir = os.path.join(ivr_pasta,prompt)
            shutil.copy(prompt_path, new_dir)
            os.remove(prompt_path)
            prompt_name = prefixo_prompt + prompt[0:prompt.index(".")] + sufixo_prompt
            prompts = api_genesys.architect_api.get_architect_prompts(name=prompt_name)
            user_prompt = prompts.entities[0]
            user_prompt_resource = api_genesys.architect_api.get_architect_prompt_resource(user_prompt.id, language)
            api_genesys.post_architect_prompt_upload(user_prompt_resource.upload_uri, prompt  + '.wav', new_dir)
            success, msg = True, 'Prompt gravado com sucesso'
        except Exception as erro:
            success, msg = False, str(erro)
        finally:
            ws.append([prompt_name, ivr, prompt, success, msg])

    wb.save(f'{ivr_pasta}/dados.xlsx')


def upload_user_prompts(api_genesys: Genesys, diretorio: str, ivr: str, language:str = 'pt-br', prefixo_prompt: str = '', sufixo_prompt: str = ''):
    """
    Atualização e criação de prompts, pelos arquivos de prompts no diretorio
    """
    lista_prompts = os.listdir(diretorio)
    wb = Workbook()
    ws = wb.active
    ws.append(['Nome', 'IVR', 'Prompt', 'Sucesso', 'Message'])
    ivr_pasta = f'prompts/{ivr}/'
    os.mkdir(ivr_pasta)
    for prompt in lista_prompts:
        prompt_path = os.path.join(diretorio,prompt)
        if not os.path.isfile(prompt_path):
            continue
        try:
            new_dir = os.path.join(ivr_pasta,prompt)
            shutil.copy(prompt_path, new_dir)
            os.remove(prompt_path)
            prompt_name = prefixo_prompt + prompt[0:prompt.index(".")] + sufixo_prompt
            prompts = api_genesys.architect_api.get_architect_prompts(name=prompt_name)
            user_prompt = prompts.entities[0]
            user_prompt_resource = api_genesys.architect_api.get_architect_prompt_resource(user_prompt.id, language)
            api_genesys.post_architect_prompt_upload(user_prompt_resource.upload_uri, prompt  + '.wav', new_dir)
            success, msg = True, 'Prompt gravado com sucesso'
        except Exception as erro:
            success, msg = False, str(erro)
        finally:
            ws.append([prompt_name, f'IVR-{ivr}', prompt, success, msg])

    wb.save(f'{ivr_pasta}/dados.xlsx')


def upload_user_prompts_for_bot(api_genesys: Genesys, workbook: Workbook, ivr: str, qtd_linhas_dataframe: int, language:str = 'pt-br', prefixo_prompt: str = '', sufixo_prompt: str = ''):
    try:
        sheet = workbook['Prompts']
        assert 'Prompts' in workbook.sheetnames
        for prompt, conteudo in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            try:
                if not prompt is None:
                    prompt_name = prefixo_prompt + prompt + sufixo_prompt
                    user_prompt = api_genesys.post_architect_prompts(prompt_name, ivr)
                    user_prompt_resource = api_genesys.post_architect_prompt_resources(user_prompt.id, language, conteudo, conteudo)
            except Exception as error:
                print(f'Erro: {error}\n{prompt=}')
    except Exception as error:
                print(f'Erro: {error}')


api_genesys = Genesys('MOVIDA')

upload_user_prompts_for_ura3(api_genesys, r'C:\Users\souza\OneDrive\Área de Trabalho\ASSISTÊNCIA', 'IVR-210749')
#ivr_clubesaude_
#ivr_quali_
