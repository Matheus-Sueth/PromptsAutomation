from genesys.api import Genesys
import openpyxl
import openpyxl.workbook
import time
import os

api_genesys = Genesys('MOVIDA')
workbook = openpyxl.load_workbook('./docs/dados10.xlsx')
tamanho = len(os.listdir('./docs/'))

for sheetname in workbook.sheetnames:
    sheet = workbook[sheetname]
    workbook_2 = openpyxl.Workbook()
    ws = workbook_2.active
    ws.append(['Prompt','PromptId','Encontrado'])
    for row_index, (prompt, conteudo) in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True), start=2):
        if row_index % 25 == 0:
            print(f'Linha {row_index} atingida. Aguarde 60seg')
            time.sleep(60)
        try:
            if not prompt is None:
                print(prompt)
                user_prompt = api_genesys.get_user_prompt_by_name_or_description(name=prompt)
                ws.append([user_prompt.name,user_prompt.id, True])
        except Exception as error:
            print(f'Exception: {error}')
            ws.append([prompt,f'Exception: {error}', False])
    sheetname_tratado = ''.join(char for char in sheetname if sheetname.isalnum())
    workbook_2.save(f'./docs/prompts{tamanho}_{sheetname_tratado}.xlsx')
    