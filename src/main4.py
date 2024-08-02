from genesys.api import Genesys
import openpyxl
import openpyxl.workbook

api_genesys = Genesys('MOVIDA')
workbook = openpyxl.load_workbook('./docs/dados5.xlsx')
workbook_2 = openpyxl.Workbook()
ws = workbook_2.active
ws.append(['Prompt','PromptId','Encontrado'])

for sheetname in workbook.sheetnames:
    sheet = workbook[sheetname]
    for prompt, conteudo in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        try:
            if not prompt is None:
                user_prompt = api_genesys.get_user_prompt_by_name_or_description(name=prompt)
                api_genesys.update_user_prompt_resource(user_prompt.id, 'pt-br', '', conteudo)
                ws.append([user_prompt.name,user_prompt.id, True])
        except Exception as error:
            print(f'Exception: {error}')
            ws.append([user_prompt.name,user_prompt.id, False])
    workbook_2.save(f'./docs/prompts5.xlsx')
    