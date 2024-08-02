import openpyxl
from genesys.api import Genesys
import time
import os

api_genesys = Genesys('MOVIDA')
workbook = openpyxl.load_workbook('./docs/dados10.xlsx')
sheet = workbook['Sheet']
description = input('Digite o IVR: ')
tamanho = len(os.listdir('./docs/'))
workbook_2 = openpyxl.Workbook()
ws = workbook_2.active
ws.append(['Prompt','PromptId','Encontrado'])
for row_index, (name_prompt, conteudo) in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True), start=2):
    if row_index % 25 == 0:
        workbook_2.save(f'./docs/prompts{tamanho}_{row_index}.xlsx')
        print(f'Linha {row_index}: {name_prompt} atingiu o limite. Aguarde 60seg')
        time.sleep(60)
        workbook_2 = openpyxl.Workbook()
        ws = workbook_2.active
        ws.append(['Prompt','PromptId','Encontrado'])
    try:
        print(f'{name_prompt=}')
        if not name_prompt is None:
            user_prompt = api_genesys.create_new_user_prompt(name_prompt, description)
            user_prompt_resource = api_genesys.create_new_user_prompt_resource(user_prompt.id, 'pt-br', '', '')
            api_genesys.upload_user_prompt_resource_by_url(user_prompt_resource.uploadUri, 'Silence_300ms.wav', r'.\prompts\Silence_300ms.wav')
            ws.append([user_prompt.name,user_prompt.id, True])
    except Exception as error:
        print(f'Erro: {error}\n')
        ws.append([name_prompt,f'Erro: {error}', False])

