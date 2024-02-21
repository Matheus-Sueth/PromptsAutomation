import openpyxl
from genesys.api import Genesys

api_genesys = Genesys('MOVIDA')
workbook = openpyxl.load_workbook('arquivo3.xlsx')
sheet = workbook['Sheet1']
nome_bot = input('Digite a sigla do bot ou digite 1 caso não precise digitar a sigla: ')
description = input('Digite o IVR: ')

for prompt, conteudo in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    try:
        if not prompt is None:
            if nome_bot != "1":
                palavras = prompt.split("_")
                palavras.insert(-1,nome_bot)
                prompt = '_'.join(palavras)
            user_prompt = api_genesys.create_new_user_prompt(prompt, description)
            user_prompt_resource = api_genesys.create_new_user_prompt_resource(user_prompt['id'], 'pt-br', conteudo, conteudo)
    except Exception as error:
        print(f'Erro: {error}\n{prompt=}')
